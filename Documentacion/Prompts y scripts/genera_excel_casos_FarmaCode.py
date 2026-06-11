#!/usr/bin/env python3
"""
Genera el Excel de casos de prueba para FarmaCode.
Lee los JSON de Documentacion/Casos_Prueba/ y produce
Documentacion/Casos_Prueba/FarmaCode_casos_prueba.xlsx con 2 hojas: Resumen y Detalle.
"""
import json, glob, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CASOS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Documentacion", "Casos_Prueba")
OUTPUT    = os.path.join(CASOS_DIR, "FarmaCode_casos_prueba.xlsx")

HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
CELL_FONT   = Font(name="Calibri", size=10)
WRAP        = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin")
)

PRIO_COLORS = {
    "Alta":  PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "Media": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
    "Baja":  PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),
}
RIESGO_COLORS = {
    "Alto":  PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid"),
    "Medio": PatternFill(start_color="FFDD99", end_color="FFDD99", fill_type="solid"),
    "Bajo":  PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),
}

def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        c = ws.cell(row=row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = WRAP
        c.border = THIN_BORDER

def style_cell(ws, row, col, prio=None, riesgo=None):
    c = ws.cell(row=row, column=col)
    c.font = CELL_FONT
    c.alignment = WRAP
    c.border = THIN_BORDER
    if prio and prio in PRIO_COLORS:
        c.fill = PRIO_COLORS[prio]
    elif riesgo and riesgo in RIESGO_COLORS:
        c.fill = RIESGO_COLORS[riesgo]

# ── Cargar resumen consolidado ────────────────────────────────────────────────
resumen_file = os.path.join(CASOS_DIR, "casos_prueba_resumen_FarmaCode.json")
if not os.path.exists(resumen_file):
    print(f"ERROR: no se encontro {resumen_file}")
    sys.exit(1)

data = json.load(open(resumen_file, encoding="utf-8"))

# ── Cargar detalles y mapeo ID->módulo desde archivos de módulo ───────────────
# Archivos de módulo: FarmaCode_*.json (excluye casos_prueba_resumen_*)
all_detalle = []
id_to_mod   = {}

for f in sorted(glob.glob(os.path.join(CASOS_DIR, "FarmaCode_*.json"))):
    mod_data = json.load(open(f, encoding="utf-8"))
    mod_name = mod_data.get("modulo", os.path.basename(f))
    for c in mod_data.get("resumen", []):
        id_to_mod[c["id"]] = mod_name
    for c in mod_data.get("detalle", []):
        all_detalle.append(c)

detalle_map = {d["id"]: d for d in all_detalle}

wb = Workbook()

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA 1 — RESUMEN
# ═══════════════════════════════════════════════════════════════════════════════
ws_res = wb.active
ws_res.title = "Resumen"

headers_res = ["ID", "Modulo", "Area", "Nombre del Caso", "Objetivo",
               "Prioridad", "Tipo", "Riesgo", "Origen"]
for col, h in enumerate(headers_res, 1):
    ws_res.cell(row=1, column=col, value=h)
style_header(ws_res, 1, len(headers_res))

for i, caso in enumerate(data["resumen"], 2):
    modulo = id_to_mod.get(caso["id"], "")
    vals = [
        caso["id"], modulo, caso["area"], caso["nombre"], caso["objetivo"],
        caso["prioridad"], caso["tipo"], caso["riesgo"], caso["origen"]
    ]
    for col, val in enumerate(vals, 1):
        ws_res.cell(row=i, column=col, value=val)
        style_cell(ws_res, i, col,
                   prio   = caso["prioridad"] if col == 6 else None,
                   riesgo = caso["riesgo"]    if col == 8 else None)

widths_res = [8, 30, 22, 45, 65, 10, 16, 10, 18]
for col, w in enumerate(widths_res, 1):
    ws_res.column_dimensions[get_column_letter(col)].width = w
ws_res.auto_filter.ref = f"A1:{get_column_letter(len(headers_res))}{len(data['resumen'])+1}"
ws_res.freeze_panes = "A2"

# ═══════════════════════════════════════════════════════════════════════════════
# HOJA 2 — DETALLE
# ═══════════════════════════════════════════════════════════════════════════════
ws_det = wb.create_sheet("Detalle")

headers_det = ["ID", "Modulo", "Nombre del Caso", "Objetivo",
               "Precondiciones", "Pasos",
               "Datos de Prueba", "Resultado Esperado", "Criterio Aprobacion",
               "Objetos BD", "Endpoints Involucrados", "Notas Tecnicas",
               "Prioridad", "Tipo", "Riesgo"]
for col, h in enumerate(headers_det, 1):
    ws_det.cell(row=1, column=col, value=h)
style_header(ws_det, 1, len(headers_det))

row = 2
for caso in data["resumen"]:
    det     = detalle_map.get(caso["id"], {})
    modulo  = id_to_mod.get(caso["id"], "")

    precond   = "\n".join(det.get("precondiciones", []))
    pasos     = "\n".join(det.get("pasos", []))
    objetos   = "\n".join(det.get("objetos_bd_involucrados", []))
    endpoints = "\n".join(det.get("endpoints_involucrados", []))

    vals = [
        caso["id"], modulo, caso["nombre"], caso["objetivo"],
        precond, pasos,
        det.get("datos_prueba", ""),
        det.get("resultado_esperado", ""),
        det.get("criterio_aprobacion", ""),
        objetos, endpoints,
        det.get("notas_tecnicas", ""),
        caso["prioridad"], caso["tipo"], caso["riesgo"]
    ]
    for col, val in enumerate(vals, 1):
        ws_det.cell(row=row, column=col, value=val)
        style_cell(ws_det, row, col,
                   prio   = caso["prioridad"] if col == 13 else None,
                   riesgo = caso["riesgo"]    if col == 15 else None)
    ws_det.row_dimensions[row].height = 80
    row += 1

widths_det = [8, 28, 38, 55, 40, 55, 30, 45, 38, 35, 30, 45, 10, 14, 10]
for col, w in enumerate(widths_det, 1):
    ws_det.column_dimensions[get_column_letter(col)].width = w
ws_det.auto_filter.ref = f"A1:{get_column_letter(len(headers_det))}{row-1}"
ws_det.freeze_panes = "A2"

# ── Guardar ───────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"OK Excel creado: {OUTPUT}")
print(f"   Casos en Resumen : {len(data['resumen'])}")
print(f"   Casos en Detalle : {len(detalle_map)}")
print(f"   Modulos cargados : {len(id_to_mod)} IDs desde {len(sorted(glob.glob(os.path.join(CASOS_DIR, 'FarmaCode_*.json'))))} archivos")
