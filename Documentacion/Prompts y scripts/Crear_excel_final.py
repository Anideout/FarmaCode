import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Crear workbook
wb = Workbook()
wb.remove(wb.active)  # Remover hoja por defecto

# Definir estilos
header_fill = PatternFill(start_color='2D5016', end_color='2D5016', fill_type='solid')
header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
alternate_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
data_font = Font(name='Arial', size=10)
border = Border(
    left=Side(style='thin', color='000000'),
    right=Side(style='thin', color='000000'),
    top=Side(style='thin', color='000000'),
    bottom=Side(style='thin', color='000000')
)
center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

# Archivos CSV y nombres de pestañas
csv_files = [
    ('modulos_SIGISO-Sorepa.csv', 'Módulos y Funcionalidades'),
    ('roles_SIGISO-Sorepa.csv', 'Roles y Perfiles'),
    ('reglas_integraciones_SIGISO-Sorepa.csv', 'Reglas e Integraciones'),
    ('estadisticas_SIGISO-Sorepa.csv', 'Estadísticas')
]

for csv_file, sheet_name in csv_files:
    print(f'Procesando {csv_file}...')
    
    # Crear nueva hoja
    ws = wb.create_sheet(title=sheet_name)
    
    # Leer CSV
    with open(csv_file, 'r', encoding='utf-8') as f:
        reader = csv.reader(f)
        rows = list(reader)
    
    # Escribir datos
    for row_idx, row in enumerate(rows, start=1):
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value.strip() if value else '')
            
            # Aplicar estilos
            cell.border = border
            
            if row_idx == 1:  # Header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_alignment
            else:  # Data rows
                cell.font = data_font
                cell.alignment = left_alignment
                
                # Filas alternas
                if row_idx % 2 == 0:
                    cell.fill = alternate_fill
    
    # Ajustar ancho de columnas
    for col_idx in range(1, len(rows[0]) + 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        
        for row in rows:
            if col_idx <= len(row):
                cell_value = str(row[col_idx - 1])
                max_length = max(max_length, len(cell_value))
        
        # Limitar ancho máximo a 50 caracteres
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Congelar primera fila
    ws.freeze_panes = 'A2'
    
    print(f'  ✓ {sheet_name} completada')

# Guardar archivo
output_file = 'Analisis funcional_Preliminar - SIGISO-Sorepa.xlsx'
wb.save(output_file)
print(f'\n✓ Archivo Excel creado exitosamente: {output_file}')
print(f'  - 4 pestañas creadas')
print(f'  - Formato profesional aplicado')
print(f'  - Encabezados con fondo verde oscuro (#2D5016) y texto blanco')
print(f'  - Filas alternas en verde claro (#E8F5E9)')
print(f'  - Bordes en todas las celdas')
print(f'  - Fuente Arial (11pt encabezados, 10pt datos)')
print(f'  - Columnas ajustadas automáticamente')
print(f'  - Primera fila congelada')

# Made with Bob
