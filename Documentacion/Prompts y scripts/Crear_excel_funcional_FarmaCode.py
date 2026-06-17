import csv
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Documentacion", "Funcional")
OUTPUT   = os.path.join(BASE_DIR, "Levantamiento_Funcional_FarmaCode.xlsx")

wb = Workbook()
wb.remove(wb.active)

# Estilos (misma paleta que el original)
header_fill    = PatternFill(start_color='2D5016', end_color='2D5016', fill_type='solid')
header_font    = Font(name='Arial', size=11, bold=True, color='FFFFFF')
subheader_fill = PatternFill(start_color='4A7C2F', end_color='4A7C2F', fill_type='solid')
separator_fill = PatternFill(start_color='BBBBBB', end_color='BBBBBB', fill_type='solid')
separator_font = Font(name='Arial', size=10, bold=True, italic=True)
alternate_fill = PatternFill(start_color='E8F5E9', end_color='E8F5E9', fill_type='solid')
data_font      = Font(name='Arial', size=10)
border         = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)
center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align   = Alignment(horizontal='left',   vertical='center', wrap_text=True)

csv_tabs = [
    ('modulos_FarmaCode.csv',              'Pantallas'),
    ('roles_FarmaCode.csv',                'Roles'),
    ('reglas_integraciones_FarmaCode.csv', 'Reglas e Integraciones'),
    ('legacy_FarmaCode.csv',               'Hallazgos Legacy'),
    ('estadisticas_FarmaCode.csv',         'Estadísticas'),
]

for csv_name, tab_name in csv_tabs:
    path = os.path.join(BASE_DIR, csv_name)
    print(f'Procesando {csv_name}...')
    ws = wb.create_sheet(title=tab_name)

    with open(path, encoding='utf-8', newline='') as f:
        rows = list(csv.reader(f))

    out_row      = 0
    data_count   = 0
    prev_was_sep = False
    first_hdr    = False  # True después de escribir el primer encabezado

    for row in rows:
        text = ','.join(row).strip()

        # Saltar filas vacías
        if not text:
            continue

        is_sep = '---' in text
        # Primera fila no-separador = encabezado; tras un separador la siguiente también
        is_hdr = (not is_sep) and (not first_hdr or prev_was_sep)

        out_row += 1

        for col_idx, val in enumerate(row, 1):
            c = ws.cell(row=out_row, column=col_idx, value=val.strip() if val else '')
            c.border = border

            if is_sep:
                c.fill = separator_fill
                c.font = separator_font
                c.alignment = center_align
            elif is_hdr:
                # Primer encabezado: verde oscuro; sub-encabezado de sección 2: verde medio
                c.fill = header_fill if not first_hdr else subheader_fill
                c.font = header_font
                c.alignment = center_align
            else:
                c.font = data_font
                c.alignment = left_align
                if data_count % 2 == 0:
                    c.fill = alternate_fill

        if is_hdr and not first_hdr:
            first_hdr = True
        if not is_sep and not is_hdr:
            data_count += 1

        prev_was_sep = is_sep

    # Ajustar ancho de columnas (máximo 50)
    for col_idx in range(1, ws.max_column + 1):
        max_len = 0
        col_letter = get_column_letter(col_idx)
        for row_cells in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for c in row_cells:
                if c.value:
                    max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    ws.freeze_panes = 'A2'
    print(f'  OK {tab_name}: {out_row} filas escritas')

wb.save(OUTPUT)
print(f'\nOK Archivo Excel creado: {OUTPUT}')
print(f'  Hojas: {", ".join(ws.title for ws in wb.worksheets)}')
