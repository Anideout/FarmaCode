#!/usr/bin/env python3
"""
Genera un Excel por app desde los JSON de casos de prueba.
Cada Excel tiene 2 hojas: Resumen y Detalle.
Uso: python3 genera_excel_casos.py [APP_NAME]
  Sin argumento genera para todas las apps.
"""
import json, glob, sys, os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

APPS = ["AdmPublicadorSGA", "DPROC", "PAUTAS_TERRENO", "POYM", "SGA"]
DIR = os.path.dirname(os.path.abspath(__file__))

HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Calibri", size=10)
WRAP = Alignment(wrap_text=True, vertical="top")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

PRIO_COLORS = {
    "Alta": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
    "Media": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
    "Baja": PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),
}

def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
        cell.border = THIN_BORDER

def style_cell(ws, row, col, prio=None):
    cell = ws.cell(row=row, column=col)
    cell.font = CELL_FONT
    cell.alignment = WRAP
    cell.border = THIN_BORDER
    if prio and prio in PRIO_COLORS:
        cell.fill = PRIO_COLORS[prio]

def generate_excel(app_name):
    resumen_file = os.path.join(DIR, f"{app_name}_00_casos_prueba_resumen.json")
    if not os.path.exists(resumen_file):
        print(f"  ❌ {resumen_file} no encontrado")
        return

    data = json.load(open(resumen_file))
    
    # Cargar detalles desde archivos de módulo
    all_detalle = []
    for f in sorted(glob.glob(os.path.join(DIR, f"{app_name}_[01][0-9]_*.json"))):
        mod_data = json.load(open(f))
        all_detalle.extend(mod_data.get("detalle", []))
    
    detalle_map = {d["id"]: d for d in all_detalle}

    wb = Workbook()

    # === HOJA RESUMEN ===
    ws_res = wb.active
    ws_res.title = "Resumen"
    
    headers_res = ["ID", "Módulo", "Área", "Nombre", "Objetivo", "Prioridad", "Tipo", "Riesgo Migración", "Origen", "Trazabilidad"]
    for col, h in enumerate(headers_res, 1):
        ws_res.cell(row=1, column=col, value=h)
    style_header(ws_res, 1, len(headers_res))

    # Pre-cargar mapeo ID -> módulo
    id_to_mod = {}
    for f in sorted(glob.glob(os.path.join(DIR, f"{app_name}_[01][0-9]_*.json"))):
        mod_data = json.load(open(f))
        mod_name = mod_data.get("modulo", os.path.basename(f))
        for c in mod_data.get("resumen", []):
            id_to_mod[c["id"]] = mod_name

    for i, caso in enumerate(data["resumen"], 2):
        modulo = id_to_mod.get(caso["id"], "")

        traz = ", ".join(caso.get("trazabilidad", []))
        vals = [caso["id"], modulo, caso["area"], caso["nombre"], caso["objetivo"],
                caso["prioridad"], caso["tipo"], caso["riesgo_migracion"], caso["origen"], traz]
        for col, val in enumerate(vals, 1):
            ws_res.cell(row=i, column=col, value=val)
            style_cell(ws_res, i, col, prio=caso["prioridad"] if col == 6 else None)

    # Anchos de columna resumen
    widths_res = [8, 30, 18, 40, 60, 10, 14, 16, 16, 40]
    for col, w in enumerate(widths_res, 1):
        ws_res.column_dimensions[get_column_letter(col)].width = w
    ws_res.auto_filter.ref = f"A1:{get_column_letter(len(headers_res))}{len(data['resumen'])+1}"
    ws_res.freeze_panes = "A2"

    # === HOJA DETALLE ===
    ws_det = wb.create_sheet("Detalle")
    
    headers_det = ["ID", "Módulo", "Nombre", "Objetivo", "Precondiciones", "Pasos",
                   "Datos de Prueba", "Resultado Esperado", "Criterio Aprobación",
                   "Objetos BD", "SPs Involucrados", "Notas Migración",
                   "Prioridad", "Tipo", "Riesgo", "Trazabilidad"]
    for col, h in enumerate(headers_det, 1):
        ws_det.cell(row=1, column=col, value=h)
    style_header(ws_det, 1, len(headers_det))

    row = 2
    for caso in data["resumen"]:
        det = detalle_map.get(caso["id"], {})
        
        modulo = id_to_mod.get(caso["id"], "")

        precond = "\n".join(det.get("precondiciones", []))
        pasos = "\n".join(det.get("pasos", []))
        objetos = "\n".join(det.get("objetos_bd_involucrados", []))
        sps = "\n".join(det.get("sps_involucrados", []))
        traz = ", ".join(caso.get("trazabilidad", []))

        vals = [
            caso["id"], modulo, caso["nombre"], caso["objetivo"],
            precond, pasos,
            det.get("datos_prueba", ""),
            det.get("resultado_esperado", ""),
            det.get("criterio_aprobacion", ""),
            objetos, sps,
            det.get("notas_migracion", ""),
            caso["prioridad"], caso["tipo"], caso["riesgo_migracion"], traz
        ]
        for col, val in enumerate(vals, 1):
            ws_det.cell(row=row, column=col, value=val)
            style_cell(ws_det, row, col, prio=caso["prioridad"] if col == 13 else None)
        row += 1

    # Anchos de columna detalle
    widths_det = [8, 28, 35, 50, 40, 50, 30, 40, 35, 35, 30, 40, 10, 14, 10, 35]
    for col, w in enumerate(widths_det, 1):
        ws_det.column_dimensions[get_column_letter(col)].width = w
    ws_det.auto_filter.ref = f"A1:{get_column_letter(len(headers_det))}{row-1}"
    ws_det.freeze_panes = "A2"

    # Guardar
    output = os.path.join(DIR, f"{app_name}_casos_prueba.xlsx")
    wb.save(output)
    print(f"  ✅ {app_name}: {len(data['resumen'])} casos -> {os.path.basename(output)}")

if __name__ == "__main__":
    apps = [sys.argv[1]] if len(sys.argv) > 1 else APPS
    print("Generando Excel de casos de prueba...")
    for app in apps:
        generate_excel(app)
    print("Listo.")
